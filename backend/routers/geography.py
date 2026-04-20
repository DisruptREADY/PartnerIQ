"""Geography router — metro, micro, state, county, and place list endpoints."""
import logging
import time
from typing import Dict, List, Optional

from fastapi import APIRouter, Query
from fastapi.responses import JSONResponse

from ..config import CENSUS_API_KEY
from ..utils import request_with_retry

log = logging.getLogger("data_portal")
router = APIRouter()

# Simple in-memory caches
_metro_cache: Dict[str, object] = {"data": None, "timestamp": 0}
_micro_cache: Dict[str, object] = {"data": None, "timestamp": 0}
_state_cache: Dict[str, object] = {"data": None, "timestamp": 0}
_county_cache: Dict[str, object] = {"data": None, "timestamp": 0}
_place_cache: Dict[str, object] = {"data": None, "timestamp": 0}
_cbsa_county_cache: Dict[str, object] = {"data": None, "timestamp": 0}
CACHE_TTL = 3600  # 1 hour


def _discover_cbsas_from_endpoint(url: str, api_key: Optional[str] = None) -> Dict[str, str]:
    """Query a Census ACS endpoint for all CBSAs. Returns {cbsa: name}."""
    params = {
        "get": "NAME",
        "for": "metropolitan statistical area/micropolitan statistical area:*",
    }
    if api_key:
        params["key"] = api_key

    try:
        resp = request_with_retry("GET", url, params=params)
        data = resp.json()
    except Exception as e:
        log.error("Failed to discover CBSAs from %s: %s", url, e)
        return {}

    if not isinstance(data, list) or len(data) < 2:
        log.error("Unexpected CBSA discovery response shape from %s", url)
        return {}

    header, *rows = data
    try:
        cbsa_col = header.index("metropolitan statistical area/micropolitan statistical area")
        name_col = header.index("NAME")
    except ValueError as e:
        log.error("CBSA discovery missing expected column: %s", e)
        return {}

    result: Dict[str, str] = {}
    for row in rows:
        result[row[cbsa_col]] = row[name_col]
    return result


def _discover_all_cbsas(api_key: Optional[str] = None):
    """Discover metros from ACS 1-year and micros from ACS 5-year.

    ACS 1-year only covers areas with 65K+ population, which misses most
    micropolitan areas.  ACS 5-year covers all geographies.
    """
    # Metros: ACS 1-year is fine (all metros are 65K+)
    log.info("Discovering metros from ACS1 2024 ...")
    all_1yr = _discover_cbsas_from_endpoint(
        "https://api.census.gov/data/2024/acs/acs1", api_key=api_key,
    )
    metros = {k: v for k, v in all_1yr.items() if "Metro" in v}
    log.info("Found %d metros from ACS1.", len(metros))

    # Micros: ACS 5-year catches all micros (many are < 65K pop)
    log.info("Discovering micros from ACS5 2023 ...")
    all_5yr = _discover_cbsas_from_endpoint(
        "https://api.census.gov/data/2023/acs/acs5", api_key=api_key,
    )
    micros = {k: v for k, v in all_5yr.items() if "Micro" in v}
    log.info("Found %d micros from ACS5.", len(micros))

    return metros, micros


def _discover_states(year: int = 2024, api_key: Optional[str] = None) -> Dict[str, str]:
    """Query Census for all states. Returns {fips: name}."""
    url = f"https://api.census.gov/data/{year}/acs/acs1"
    params = {"get": "NAME", "for": "state:*"}
    if api_key:
        params["key"] = api_key

    log.info("Discovering states from ACS1 %d ...", year)
    try:
        resp = request_with_retry("GET", url, params=params)
        data = resp.json()
    except Exception as e:
        log.error("Failed to discover states: %s", e)
        return {}

    if not isinstance(data, list) or len(data) < 2:
        log.error("Unexpected state discovery response shape")
        return {}

    header, *rows = data
    try:
        state_col = header.index("state")
        name_col = header.index("NAME")
    except ValueError as e:
        log.error("State discovery missing expected column: %s", e)
        return {}

    states: Dict[str, str] = {}
    for row in rows:
        states[row[state_col]] = row[name_col]
    log.info("Found %d states.", len(states))
    return states


def _discover_counties(year: int = 2023, api_key: Optional[str] = None) -> Dict[str, str]:
    """Query Census ACS 5-year for all counties (covers all 3,200+ counties)."""
    url = f"https://api.census.gov/data/{year}/acs/acs5"
    params = {"get": "NAME", "for": "county:*"}
    if api_key:
        params["key"] = api_key

    log.info("Discovering counties from ACS5 %d ...", year)
    try:
        resp = request_with_retry("GET", url, params=params)
        data = resp.json()
    except Exception as e:
        log.error("Failed to discover counties: %s", e)
        return {}

    if not isinstance(data, list) or len(data) < 2:
        log.error("Unexpected county discovery response shape")
        return {}

    header, *rows = data
    try:
        state_col = header.index("state")
        county_col = header.index("county")
        name_col = header.index("NAME")
    except ValueError as e:
        log.error("County discovery missing expected column: %s", e)
        return {}

    counties: Dict[str, str] = {}
    for row in rows:
        fips = row[state_col] + row[county_col]
        counties[fips] = row[name_col]
    log.info("Found %d counties.", len(counties))
    return counties


def _discover_places(year: int = 2023, api_key: Optional[str] = None) -> Dict[str, str]:
    """Query Census ACS 5-year for all places. Returns {state_fips+place_fips: name}."""
    url = f"https://api.census.gov/data/{year}/acs/acs5"
    params = {"get": "NAME", "for": "place:*"}
    if api_key:
        params["key"] = api_key

    log.info("Discovering places from ACS5 %d ...", year)
    try:
        resp = request_with_retry("GET", url, params=params)
        data = resp.json()
    except Exception as e:
        log.error("Failed to discover places: %s", e)
        return {}

    if not isinstance(data, list) or len(data) < 2:
        log.error("Unexpected place discovery response shape")
        return {}

    header, *rows = data
    try:
        state_col = header.index("state")
        place_col = header.index("place")
        name_col = header.index("NAME")
    except ValueError as e:
        log.error("Place discovery missing expected column: %s", e)
        return {}

    places: Dict[str, str] = {}
    for row in rows:
        fips = row[state_col] + row[place_col]
        places[fips] = row[name_col]
    log.info("Found %d places.", len(places))
    return places


def _ensure_cbsa_caches():
    """Populate both metro and micro caches from a single API call."""
    now = time.time()
    metro_stale = _metro_cache["data"] is None or (now - _metro_cache["timestamp"]) >= CACHE_TTL
    micro_stale = _micro_cache["data"] is None or (now - _micro_cache["timestamp"]) >= CACHE_TTL

    if metro_stale or micro_stale:
        metros, micros = _discover_all_cbsas(api_key=CENSUS_API_KEY)
        if metros:
            _metro_cache["data"] = metros
            _metro_cache["timestamp"] = now
        if micros:
            _micro_cache["data"] = micros
            _micro_cache["timestamp"] = now


def _discover_cbsa_county_map(year: int = 2023, api_key: Optional[str] = None) -> Dict[str, List[str]]:
    """Discover which counties belong to each CBSA (metro and micro).

    Downloads the Census Bureau's official CBSA delineation file (Excel) which
    lists every county and its assigned CBSA code. This replaced the previous
    ACS geography hierarchy query which Census discontinued.
    """
    import io
    import openpyxl

    url = (
        f"https://www2.census.gov/programs-surveys/metro-micro/"
        f"geographies/reference-files/{year}/delineation-files/list1_{year}.xlsx"
    )
    log.info("Downloading CBSA delineation file for %d ...", year)
    try:
        resp = request_with_retry("GET", url)
    except Exception as e:
        log.error("Failed to download CBSA delineation file: %s", e)
        return {}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True)
        ws = wb.active
    except Exception as e:
        log.error("Failed to parse CBSA delineation Excel: %s", e)
        return {}

    # Find header row (contains "CBSA Code")
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        vals = [c.value for c in row]
        if "CBSA Code" in vals:
            header_row = row[0].row
            headers = vals
            break

    if header_row is None:
        log.error("Could not find header row in delineation file")
        return {}

    try:
        cbsa_idx = headers.index("CBSA Code")
        state_idx = headers.index("FIPS State Code")
        county_idx = headers.index("FIPS County Code")
    except ValueError as e:
        log.error("Delineation file missing expected column: %s", e)
        return {}

    mapping: Dict[str, List[str]] = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        cbsa_code = row[cbsa_idx]
        state_fips = row[state_idx]
        county_fips_part = row[county_idx]

        if not cbsa_code or not state_fips or not county_fips_part:
            continue

        cbsa = str(cbsa_code).strip()
        st = str(state_fips).strip().zfill(2)
        co = str(county_fips_part).strip().zfill(3)
        full_fips = st + co
        mapping.setdefault(cbsa, []).append(full_fips)

    wb.close()
    log.info("Built CBSA → county mapping: %d CBSAs, %d total counties.",
             len(mapping), sum(len(v) for v in mapping.values()))
    return mapping


def get_cbsa_county_map() -> Dict[str, List[str]]:
    """Get CBSA → county FIPS mapping with caching."""
    now = time.time()
    if _cbsa_county_cache["data"] is not None and (now - _cbsa_county_cache["timestamp"]) < CACHE_TTL:
        return _cbsa_county_cache["data"]

    mapping = _discover_cbsa_county_map(api_key=CENSUS_API_KEY)
    if mapping:
        _cbsa_county_cache["data"] = mapping
        _cbsa_county_cache["timestamp"] = now
    return mapping or {}


def get_metros() -> Dict[str, str]:
    """Get metro list with caching."""
    _ensure_cbsa_caches()
    return _metro_cache["data"] or {}


def get_micros() -> Dict[str, str]:
    """Get micro list with caching."""
    _ensure_cbsa_caches()
    return _micro_cache["data"] or {}


def get_states() -> Dict[str, str]:
    """Get state list with caching."""
    now = time.time()
    if _state_cache["data"] is not None and (now - _state_cache["timestamp"]) < CACHE_TTL:
        return _state_cache["data"]

    states = _discover_states(api_key=CENSUS_API_KEY)
    if states:
        _state_cache["data"] = states
        _state_cache["timestamp"] = now
    return states or {}


def get_counties() -> Dict[str, str]:
    """Get county list with caching."""
    now = time.time()
    if _county_cache["data"] is not None and (now - _county_cache["timestamp"]) < CACHE_TTL:
        return _county_cache["data"]

    counties = _discover_counties(api_key=CENSUS_API_KEY)
    if counties:
        _county_cache["data"] = counties
        _county_cache["timestamp"] = now
    return counties or {}


def get_places(state_fips: Optional[str] = None) -> Dict[str, str]:
    """Get places list with caching. Optionally filter by state FIPS."""
    now = time.time()
    if _place_cache["data"] is None or (now - _place_cache["timestamp"]) >= CACHE_TTL:
        places = _discover_places(api_key=CENSUS_API_KEY)
        if places:
            _place_cache["data"] = places
            _place_cache["timestamp"] = now

    all_places = _place_cache["data"] or {}

    if state_fips:
        return {k: v for k, v in all_places.items() if k[:2] == state_fips}
    return all_places


@router.get("/geographies")
async def list_geographies(
    type: str = Query("msa", pattern="^(msa|micro|county|state|place)$"),
    state: Optional[str] = Query(None),
):
    """Return list of geographies: [{id, name, type}, ...]."""
    cache_headers = {"Cache-Control": "public, max-age=3600"}
    if type == "msa":
        geos = get_metros()
        return JSONResponse(content=[
            {"id": cbsa, "name": name, "type": "msa"}
            for cbsa, name in sorted(geos.items(), key=lambda x: x[1])
        ], headers=cache_headers)
    elif type == "micro":
        geos = get_micros()
        return JSONResponse(content=[
            {"id": cbsa, "name": name, "type": "micro"}
            for cbsa, name in sorted(geos.items(), key=lambda x: x[1])
        ], headers=cache_headers)
    elif type == "state":
        geos = get_states()
        return JSONResponse(content=[
            {"id": fips, "name": name, "type": "state"}
            for fips, name in sorted(geos.items(), key=lambda x: x[1])
        ], headers=cache_headers)
    elif type == "county":
        geos = get_counties()
        if state:
            geos = {k: v for k, v in geos.items() if k[:2] == state}
        return JSONResponse(content=[
            {"id": fips, "name": name, "type": "county"}
            for fips, name in sorted(geos.items(), key=lambda x: x[1])
        ], headers=cache_headers)
    elif type == "place":
        geos = get_places(state_fips=state)
        return JSONResponse(content=[
            {"id": fips, "name": name, "type": "place"}
            for fips, name in sorted(geos.items(), key=lambda x: x[1])
        ], headers=cache_headers)
