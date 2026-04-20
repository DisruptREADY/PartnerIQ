"""Export router — Excel download endpoint."""
import io
import logging
from typing import List, Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from ..config import INDICATORS
from ..cache import get_cached

log = logging.getLogger("data_portal")
router = APIRouter()


class ExportRequest(BaseModel):
    cbsas: List[str] = []
    geo_ids: List[str] = []
    indicators: List[str]
    years: List[int]
    geo_type: str = "msa"
    acs_dataset: str = "acs1"
    coli_adjust: bool = False
    inflation_adjust: bool = False
    inflation_base_year: Optional[int] = None


@router.post("/export/xlsx")
async def export_xlsx(req: ExportRequest):
    """Build and return an Excel workbook from cached data or raise 400."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed on the server.")

    geo_ids = req.geo_ids or req.cbsas
    if not geo_ids or not req.indicators or not req.years:
        raise HTTPException(400, "geo_ids/cbsas, indicators, and years are required.")

    # Build cache key matching data router
    cache_extra = ""
    if req.coli_adjust:
        cache_extra += "_coli"
    if req.inflation_adjust and req.inflation_base_year:
        cache_extra += f"_infl{req.inflation_base_year}"

    # Try to use cached data
    cached = get_cached(geo_ids, req.indicators, req.years, req.geo_type,
                        req.acs_dataset + cache_extra)
    if not cached:
        raise HTTPException(
            400,
            "No cached data available. Fetch data first via /api/data, then export.",
        )

    columns = cached.get("columns", [])
    rows = cached.get("rows", [])
    yoy = cached.get("yoy", {})

    # Guard against excessively large exports that could exhaust memory
    if len(rows) > 50_000:
        raise HTTPException(400, f"Export too large ({len(rows):,} rows). Narrow your selection to fewer geographies, indicators, or years.")

    wb = openpyxl.Workbook()

    # --- Data sheet ---
    ws = wb.active
    ws.title = "Data"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="0E78BE", end_color="0E78BE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Write headers
    for c_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col.get("label", col.get("key", "")))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Write data rows
    for r_idx, row in enumerate(rows, 2):
        for c_idx, col in enumerate(columns, 1):
            val = row.get(col.get("key", ""))
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except (TypeError, ValueError):
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

    ws.auto_filter.ref = ws.dimensions

    # --- YoY sheet ---
    if yoy:
        ws_yoy = wb.create_sheet("Year-over-Year Changes")
        geo_id_label = {"state": "FIPS", "county": "FIPS", "place": "FIPS", "micro": "CBSA"}.get(req.geo_type, "CBSA")
        geo_name_label = {"state": "State", "county": "County", "place": "Place", "micro": "Micro"}.get(req.geo_type, "Metro")
        yoy_headers = [geo_id_label, geo_name_label, "Indicator", "Year", "Prior Year", "Change", "Change Type"]
        for c_idx, h in enumerate(yoy_headers, 1):
            cell = ws_yoy.cell(row=1, column=c_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Build CBSA->Metro lookup from rows
        cbsa_to_metro = {}
        for r in rows:
            if r.get("CBSA") and r.get("Metro"):
                cbsa_to_metro[r["CBSA"]] = r["Metro"]

        r_idx = 2
        for ind_key, changes in yoy.items():
            meta = INDICATORS.get(ind_key, {})
            label = meta.get("label", ind_key)
            for ch in changes:
                ws_yoy.cell(row=r_idx, column=1, value=ch.get("cbsa", "")).border = thin_border
                ws_yoy.cell(row=r_idx, column=2, value=cbsa_to_metro.get(ch.get("cbsa", ""), "")).border = thin_border
                ws_yoy.cell(row=r_idx, column=3, value=label).border = thin_border
                ws_yoy.cell(row=r_idx, column=4, value=ch.get("year")).border = thin_border
                ws_yoy.cell(row=r_idx, column=5, value=ch.get("prior_year")).border = thin_border

                change_cell = ws_yoy.cell(row=r_idx, column=6, value=ch.get("change"))
                change_cell.border = thin_border
                # Color-code: green for good, red for bad
                higher_is = meta.get("higher_is", "neutral")
                change_val = ch.get("change", 0)
                if higher_is == "better" and change_val > 0:
                    change_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif higher_is == "better" and change_val < 0:
                    change_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif higher_is == "worse" and change_val < 0:
                    change_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif higher_is == "worse" and change_val > 0:
                    change_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                ws_yoy.cell(row=r_idx, column=7, value=ch.get("change_type", "")).border = thin_border
                r_idx += 1

        for col_cells in ws_yoy.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except (TypeError, ValueError):
                    pass
            ws_yoy.column_dimensions[col_letter].width = min(max_len + 3, 30)

    # --- Notes sheet ---
    ws_notes = wb.create_sheet("Notes")
    acs_label = "5-Year" if req.acs_dataset == "acs5" else "1-Year"

    # Determine which sources are actually used in this export
    used_sources = set()
    for ind_key in req.indicators:
        meta = INDICATORS.get(ind_key, {})
        used_sources.add(meta.get("source", ""))

    notes = [
        ["Data Portal Export"],
        [""],
        ["Sources:"],
    ]
    if "census" in used_sources or "derived" in used_sources:
        notes.append([f"  Census ACS {acs_label} Estimates — api.census.gov"])
    if "bls" in used_sources:
        notes.append(["  Bureau of Labor Statistics (CES/LAUS) — api.bls.gov"])
    if "bea" in used_sources:
        notes.append(["  Bureau of Economic Analysis — apps.bea.gov"])
    if "fred" in used_sources:
        notes.append(["  FRED (Federal Reserve Economic Data) — fred.stlouisfed.org"])
    if "pep" in used_sources:
        notes.append(["  Census Population Estimates Program (PEP)"])
    if "qcew" in used_sources:
        notes.append(["  BLS Quarterly Census of Employment and Wages (QCEW) — data.bls.gov/cew"])
    if "coli" in used_sources:
        notes.append(["  Council for Community and Economic Research (C2ER) Cost of Living Index"])

    notes.append([""])
    notes.append(["Notes:"])

    if "census" in used_sources and req.acs_dataset != "acs5":
        notes.append(["  ACS 1-year estimates are unavailable for 2020 due to COVID-19 data collection disruption."])
    if "bls" in used_sources:
        bls_ind_keys = [k for k in req.indicators if INDICATORS.get(k, {}).get("source") == "bls"]
        has_ces = any(k.startswith("ind_") or k == "nonfarm_jobs" for k in bls_ind_keys)
        if has_ces:
            notes.append(["  BLS CES employment figures are in thousands. M13 = annual average; latest month used when annual average is unavailable."])
            notes.append(["  For metros where BLS does not publish separate Mining & Logging and Construction series, the combined Mining, Logging & Construction figure is used for Construction."])
    notes.append(["  YoY changes: 'pct' = percentage change, 'pp' = percentage-point change."])
    if req.coli_adjust:
        notes.append(["  COLI adjustment applied: monetary values divided by COLI composite / 100 (Q3 2025 data)."])
    if req.inflation_adjust and req.inflation_base_year:
        notes.append([f"  Inflation adjustment applied: monetary values converted to {req.inflation_base_year} constant dollars using CPI-U."])
    for r_idx, row in enumerate(notes, 1):
        for c_idx, val in enumerate(row, 1):
            ws_notes.cell(row=r_idx, column=c_idx, value=val)
    ws_notes.column_dimensions["A"].width = 70

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=data_portal_export.xlsx"},
    )
